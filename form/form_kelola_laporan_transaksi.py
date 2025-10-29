from flet import *
import mysql.connector
import datetime
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
import webbrowser

# Setup koneksi ke database MySQL
koneksi_db = mysql.connector.connect(host = "localhost", user = "root", password = "", database = "mad_uas_db_2024_202153019")
cursor = koneksi_db.cursor()

# Fungsi untuk membuat koneksi ke database MySQL
def get_db_connection():
    # Membuat koneksi baru setiap kali dipanggil
    koneksi_db = mysql.connector.connect(
        host="localhost", 
        user="root", 
        password="", 
        database="mad_uas_db_2024_202153019"
    )
    return koneksi_db

# Fungsi untuk mengambil data dari tabel tertentu
def get_data_from_table(query):
    koneksi_db = get_db_connection()  # Mendapatkan koneksi baru
    cursor = koneksi_db.cursor(dictionary=True)
    cursor.execute(query)
    data = cursor.fetchall()
    cursor.close()
    koneksi_db.close()  # Menutup koneksi setelah query selesai
    return data

def simpan_ke_excel(data_transaksi, file_name="laporan_transaksi.xlsx"):
    # Membuat DataFrame dari data transaksi
    df = pd.DataFrame(data_transaksi)

    # Menyimpan DataFrame ke file Excel
    df.to_excel(file_name, index=False, engine='openpyxl')

    # Membuka file Excel yang sudah dibuat
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # Menyesuaikan lebar kolom dengan panjang data
    for col_num in range(1, len(df.columns) + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for row_num in range(1, len(df) + 2):  # +2 untuk header dan data
            cell_value = str(ws[f"{column}{row_num}"].value)
            max_length = max(max_length, len(cell_value))
        adjusted_width = max_length + 2  # Menambah sedikit ruang untuk padding
        ws.column_dimensions[column].width = adjusted_width

    # Menyimpan file dengan lebar kolom yang sudah disesuaikan
    wb.save(file_name)

    # Kembalikan nama file untuk referensi
    return file_name

def format_tanggal_dmy(tanggal):
    return tanggal.strftime("%d-%m-%Y")

def buat_pdf_pratinjau(data_transaksi):
    pdf_file = "daftar_penjualan_warmindo.pdf"
    c = canvas.Canvas(pdf_file, pagesize=landscape(A4))  # Landscape orientation
    width, height = landscape(A4)

    def draw_header():
        # Tambahkan ikon di header
        logo_laporan = "checklist.png"  # Ganti dengan path ikon Anda
        if os.path.exists(logo_laporan):
            c.drawImage(logo_laporan, 30, height - 60, width=40, height=40, preserveAspectRatio=True, mask='auto')
        c.setFont("Helvetica-Bold", 16)
        c.drawString(80, height - 40, "Laporan Transaksi Warmindo")
        c.setFont("Helvetica", 10)
        c.drawString(80, height - 55, f"Tanggal Cetak: {format_tanggal_dmy(datetime.date.today())}")

        # Menampilkan total pendapatan di sebelah kanan Tanggal
        total_pendapatan = sum(row['dt.Subtotal'] for row in data_transaksi)
        c.drawString(width - 200, height - 55, f"Pendapatan: Rp {total_pendapatan:,.0f}")

        c.setFont("Helvetica", 10)
        c.drawString(30, height - 70, "-" * 235)

        # Header kolom
        y_header = height - 100
        c.setFont("Helvetica-Bold", 10)
        column_titles = ["No.", "ID Transaksi", "Pelanggan", "Tanggal", "Menu", "Harga", "Jumlah", "Subtotal", "Pembayaran", "Jenis"]
        column_positions = [30, 55, 130, 230, 300, 450, 510, 560, 640, 720]

        for i, title in enumerate(column_titles):
            c.drawString(column_positions[i], y_header, title)
        c.line(30, y_header - 5, width - 30, y_header - 5)  # Garis bawah header

    draw_header()
    y = height - 120
    c.setFont("Helvetica", 10)

    page_started = False  # Flag untuk menandakan jika halaman baru dimulai

    for idx, row in enumerate(data_transaksi, start=1):
        if y < 50:  # Cek apakah masih ada ruang untuk baris data
            c.showPage()  # Menambah halaman baru
            draw_header()  # Gambar ulang header di halaman baru
            y = height - 120
            page_started = True  # Halaman baru telah dimulai

        # Menentukan font yang akan digunakan
        if page_started:
            c.setFont("Helvetica", 10)  # Gunakan font biasa di halaman kedua dan seterusnya
        else:
            c.setFont("Helvetica", 10)  # Font biasa untuk halaman pertama

        # Format tanggal transaksi
        tanggal_transaksi = row['t.Tanggal_Transaksi'].date().strftime('%d-%m-%Y') if isinstance(row['t.Tanggal_Transaksi'], datetime.datetime) else str(row['t.Tanggal_Transaksi'])

        # Menampilkan konten tabel
        c.drawString(30, y, str(idx))  # Nomor
        c.drawString(55, y, str(row['t.ID_Transaksi']))  # Pelanggan
        c.drawString(130, y, row['p.Nama_Pelanggan'][:40])  # Pelanggan
        c.drawString(230, y, tanggal_transaksi)  # Tanggal
        c.drawString(300, y, row['m.Nama_Menu'][:35])  # Menu
        c.drawString(450, y, f"Rp {row['m.Harga']:,}")  # Harga
        # Menampilkan 'Jumlah' secara rata tengah di kolom yang telah ditentukan
        jumlah = str(row['dt.Jumlah'])
        jumlah_width = c.stringWidth(jumlah, "Helvetica", 10)  # Menghitung panjang teks
        kolom_jumlah_pos_x = 510  # Posisi X awal untuk kolom 'Jumlah'
        kolom_jumlah_width = 30  # Lebar kolom 'Jumlah' yang didapat dari perbedaan posisi kolom 'Jumlah' dan 'Subtotal'

        # Menghitung posisi tengah kolom 'Jumlah'
        jumlah_pos_x = kolom_jumlah_pos_x + (kolom_jumlah_width - jumlah_width) / 2  # Menghitung posisi X agar teks berada di tengah kolom
        c.setFont("Helvetica", 10)
        c.drawString(jumlah_pos_x, y, jumlah)  # Menampilkan teks 'Jumlah' pada posisi tengah kolom

        c.drawString(560, y, f"Rp {row['dt.Subtotal']:,}")  # Subtotal
        c.drawString(640, y, row['t.Metode_Pembayaran'])  # Pembayaran
        c.drawString(720, y, row['t.Jenis_Pesanan'])  # Jenis Pesanan
        y -= 20

    # Menambahkan garis horizontal sebelum keterangan Total Pendapatan
    c.setLineWidth(0.5)  # Set line width
    c.line(30, y, width - 30, y)  # Garis horizontal di bagian bawah laporan
    y -= 20  # Memberi sedikit ruang setelah garis

    # Menambahkan keterangan Pendapatan di bagian bawah laporan
    total_pendapatan = sum(row['dt.Subtotal'] for row in data_transaksi)
    c.setFont("Helvetica", 12)
    c.drawString(width - 200, y, f"Total Pendapatan: Rp {total_pendapatan:,.0f}")
    y -= 20

    c.save()
    return pdf_file

def ambil_data_laporan_transaksi(tanggal_awal=None, tanggal_akhir=None, id_pelanggan=None, status=None, jenis=None):
    # Query untuk mengambil data transaksi berdasarkan filter dengan join ke tabel pelanggan, detail_transaksi, kasir, dan menu
    query = """
        SELECT 
            dt.ID_Detail, 
            p.Nama_Pelanggan, 
            t.Tanggal_Transaksi, 
            m.Nama_Menu, 
            m.Harga, 
            dt.Jumlah, 
            dt.Subtotal, 
            t.Metode_Pembayaran,
            t.Jenis_Pesanan,
            t.ID_Transaksi
        FROM transaksi t
        JOIN detail_transaksi dt ON t.ID_Transaksi = dt.ID_Transaksi
        JOIN menu m ON dt.ID_Menu = m.ID_Menu
        JOIN pelanggan p ON t.ID_Pelanggan = p.ID_Pelanggan
    """
    
    conditions = []
    params = []

    # Menambahkan kondisi berdasarkan filter yang ada
    if tanggal_awal and tanggal_akhir:
        conditions.append("DATE(t.Tanggal_Transaksi) BETWEEN %s AND %s")
        params.extend([tanggal_awal, tanggal_akhir])
    if id_pelanggan:
        conditions.append("t.ID_Pelanggan = %s")
        params.append(id_pelanggan)
    if status:
        conditions.append("t.Metode_Pembayaran = %s")
        params.append(status)
    if jenis:
        conditions.append("t.Jenis_Pesanan = %s")
        params.append(jenis)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    
    query += " ORDER BY t.Tanggal_Transaksi DESC"
    
    # Eksekusi query
    cursor.execute(query, tuple(params))
    return cursor.fetchall()

# Halaman kelola laporan transaksi
def form_kelola_laporan_transaksi():
    # Variabel untuk paginasi
    baris_data_per_hal = 5
    hal_sekarang = 0  # Mulai dari halaman pertama

    # Input filter
    inputan_tanggal_awal = TextField(label="Tanggal Awal", width=300, hint_text = "masukkan tanggal awal ... ", expand = True, read_only=True)
    inputan_tanggal_akhir = TextField(label="Tanggal Akhir", width=300, hint_text = "masukkan tanggal akhir ... ", expand = True, read_only=True)
    # Mengambil data pelanggan dari database
    pelanggan_query = "SELECT * FROM pelanggan ORDER BY ID_Pelanggan DESC"
    pelanggan_data = get_data_from_table(pelanggan_query)  # Mendapatkan data pelanggan terbaru
    inputan_id_pelanggan = Dropdown(
        label="Nama Pelanggan", 
        width=500,
        options=[dropdown.Option(row['ID_Pelanggan'], str(row['Nama_Pelanggan'])) for row in pelanggan_data]
    )
    inputan_status_transaksi = Dropdown(label="Metode Pembayaran", width=500,
        options=[dropdown.Option('Tunai', 'Tunai'),
                 dropdown.Option('Kartu Kredit', 'Kartu Kredit'),
                 dropdown.Option('Transfer', 'Transfer')])
    inputan_jenis_transaksi = Dropdown(label="Jenis Pesanan", width=500,
        options=[dropdown.Option('Take Away', 'Take Away'),
                 dropdown.Option('Dine In', 'Dine In')])
    inputan_pencarian = TextField(label="Pencarian", width=300)

    def change_date(e):
            tgl_baru = inputan_tanggal_awal_pick.value
            inputan_tanggal_awal.value = tgl_baru.date()
            inputan_tanggal_awal.update()

    def date_picker_dismissed(e):
            tgl_baru = inputan_tanggal_awal.value
            inputan_tanggal_awal.value = tgl_baru
            inputan_tanggal_awal.update()

    inputan_tanggal_awal_pick = DatePicker(
        on_change=change_date,
        on_dismiss=date_picker_dismissed,
        first_date=datetime.datetime(1945, 1, 1),
        last_date=datetime.date.today(),
    )

    def change_date(e):
            tgl_baru = inputan_tanggal_akhir_pick.value
            inputan_tanggal_akhir.value = tgl_baru.date()
            inputan_tanggal_akhir.update()

    def date_picker_dismissed(e):
            tgl_baru = inputan_tanggal_akhir.value
            inputan_tanggal_akhir.value = tgl_baru
            inputan_tanggal_akhir.update()

    inputan_tanggal_akhir_pick = DatePicker(
        on_change=change_date,
        on_dismiss=date_picker_dismissed,
        first_date=datetime.datetime(1945, 1, 1),
        last_date=datetime.date.today(),
    )

    snack_bar_berhasil = SnackBar(Text("Operasi berhasil"), bgcolor="green")
    snack_bar_gagal = SnackBar(Text("Operasi gagal"), bgcolor="red")

    # Fungsi untuk mengupdate data laporan berdasarkan filter
    def update_laporan_transaksi(e):
        tanggal_awal = inputan_tanggal_awal.value
        tanggal_akhir = inputan_tanggal_akhir.value
        id_pelanggan = inputan_id_pelanggan.value
        status = inputan_status_transaksi.value
        jenis = inputan_jenis_transaksi.value

        data_laporan = ambil_data_laporan_transaksi(tanggal_awal, tanggal_akhir, id_pelanggan, status, jenis)

        nonlocal filtered_data_laporan
        filtered_data_laporan = data_laporan
        update_baris_data_laporan()

    # Variabel untuk menyimpan data laporan yang telah difilter
    baris_data_laporan = []  # Dideklarasikan sebelum digunakan

    # Fungsi untuk mengupdate baris data laporan
    def update_baris_data_laporan():
        nonlocal baris_data_laporan  # Gunakan nonlocal karena baris_data_laporan ada di tingkat luar
        index_mulai = hal_sekarang * baris_data_per_hal
        index_selesai = index_mulai + baris_data_per_hal
        hal_data = filtered_data_laporan[index_mulai:index_selesai]
        
        baris_data_laporan = [
            DataRow(
                cells=[
                    DataCell(Text(str(index_mulai + i + 1))),
                    DataCell(Text(str(laporan_kolom[9]))),  # ID Transaksi
                    DataCell(Text(str(laporan_kolom[1]))),  # ID Pelanggan
                    DataCell(Text(str(laporan_kolom[2].date()))),  # Tanggal Transaksi tanpa waktu
                    DataCell(Text(str(laporan_kolom[3]))),  # Jumlah
                    DataCell(Text(str(laporan_kolom[4]))),  # Status Transaksi
                    DataCell(Text(str(laporan_kolom[5]))),  
                    DataCell(Text(str(laporan_kolom[6]))),  
                    DataCell(Text(str(laporan_kolom[7]))),  
                    DataCell(Text(str(laporan_kolom[8]))),  
                ]
            )
            for i, laporan_kolom in enumerate(hal_data)
        ]
        
        tabel_data_laporan.rows = baris_data_laporan
        tabel_data_laporan.update()  # Menyegarkan DataTable untuk menampilkan baris yang sudah diperbarui

    # Fungsi untuk memfilter data berdasarkan pencarian
    def filter_menu(e):
        query_pencarian = inputan_pencarian.value.lower()
        nonlocal filtered_data_laporan
        filtered_data_laporan = [laporan for laporan in filtered_data_laporan if query_pencarian in str(laporan[0]).lower() or query_pencarian in str(laporan[1]).lower() or query_pencarian in str(laporan[4]).lower()]
        update_baris_data_laporan()

    inputan_pencarian.on_change = filter_menu

    # Variabel untuk menyimpan data laporan yang telah difilter
    filtered_data_laporan = []

    # Fungsi untuk navigasi halaman laporan
    def pergi_hal_sebelumnya(e):
        nonlocal hal_sekarang
        if hal_sekarang > 0:
            hal_sekarang -= 1
            update_baris_data_laporan()

    def pergi_hal_selanjutnya(e):
        nonlocal hal_sekarang
        if (hal_sekarang + 1) * baris_data_per_hal < len(filtered_data_laporan):
            hal_sekarang += 1
            update_baris_data_laporan()

    btn_sebelumnya = ElevatedButton("Sebelumnya", on_click=pergi_hal_sebelumnya)
    btn_selanjutnya = ElevatedButton("Berikutnya", on_click=pergi_hal_selanjutnya)

    # Tabel untuk menampilkan laporan transaksi
    tabel_data_laporan = DataTable(
        columns=[
            DataColumn(Text("No.")),
            DataColumn(Text("ID Transaksi")),
            DataColumn(Text("Nama Pelanggan")),
            DataColumn(Text("Tanggal Transaksi")),
            DataColumn(Text("Nama Menu")),
            DataColumn(Text("Harga")),
            DataColumn(Text("Jumlah")),
            DataColumn(Text("Subtotal")),
            DataColumn(Text("Metode Pembayaran")),
            DataColumn(Text("Jenis Pesanan")),
        ],
        rows=[],
        width="auto",
    )

    def pratinjau_pdf(e=None):
        # Pastikan ada data transaksi yang telah difilter
        if filtered_data_laporan:
            # Mengambil data transaksi dari filtered_data_laporan
            data_transaksi = []
            for row in filtered_data_laporan:
                # Ambil kolom data yang relevan dari tabel (contoh: nama pelanggan, tanggal transaksi, dll)
                data_transaksi.append({
                    'p.Nama_Pelanggan': row[1],  # Nama Pelanggan
                    't.Tanggal_Transaksi': row[2],  # Tanggal Transaksi
                    'm.Nama_Menu': row[3],  # Nama Menu
                    'm.Harga': row[4],  # Harga
                    'dt.Jumlah': row[5],  # Jumlah
                    'dt.Subtotal': row[6],  # Subtotal
                    't.Metode_Pembayaran': row[7],  # Metode Pembayaran
                    't.Jenis_Pesanan': row[8],  # Metode Pembayaran
                    't.ID_Transaksi': row[9],  # Metode Pembayaran
                })
            
            # Membuat file PDF pratinjau dengan data transaksi
            pdf_file = buat_pdf_pratinjau(data_transaksi)
            
            # Buka file PDF di browser
            webbrowser.open("file://" + os.path.abspath(pdf_file))
        else:
            # Jika tidak ada data transaksi, tampilkan pesan error
            snack_bar_gagal.text = "Tidak ada data yang ditemukan untuk laporan ini."
            snack_bar_gagal.open()

    def simpan_excel(e=None):
        # Pastikan ada data transaksi yang telah difilter
        if filtered_data_laporan:
            # Mengambil data transaksi dari filtered_data_laporan
            data_transaksi = []
            for row in filtered_data_laporan:
                data_transaksi.append({
                    'ID Transaksi': row[9],  # ID Transaksi
                    'Nama Pelanggan': row[1],  # Nama Pelanggan
                    'Tanggal Transaksi': row[2],  # Tanggal Transaksi
                    'Nama Menu': row[3],  # Nama Menu
                    'Harga': row[4],  # Harga
                    'Jumlah': row[5],  # Jumlah
                    'Subtotal': row[6],  # Subtotal
                    'Metode Pembayaran': row[7],  # Metode Pembayaran
                    'Jenis Pesanan': row[8],  # Jenis Pesanan
                })
            
            # Menyimpan data ke file Excel
            excel_file = simpan_ke_excel(data_transaksi)
        else:
            # Jika tidak ada data transaksi, tampilkan pesan error
            snack_bar_gagal.text = "Tidak ada data yang ditemukan untuk laporan ini."
            snack_bar_gagal.open()

    # Membuat form bagian kanan dengan search field dan table
    form_kanan = Container(
        content = Column(
            controls = [
                Text("Tabel Data", size = 14, weight = FontWeight.BOLD),
                inputan_pencarian,  # Menambahkan search field untuk pencarian
                Row(
                    controls = [
                        tabel_data_laporan  # Menampilkan DataTable
                    ],
                    scroll = "auto",  # Membolehkan scroll horizontal pada Row
                ),
                Row([btn_sebelumnya, btn_selanjutnya]),  # Menambahkan tombol navigasi halaman
            ],
            width="700",  # Lebar bagian kanan
            scroll="always",  # Membolehkan scroll vertikal pada Column
        ),
        border_radius = 20,
        border = border.all(color = colors.GREY_900, width = 0.5),
        padding = 20,
    )

    # Form untuk filter laporan transaksi
    form_filter = Container(
        Column(
            controls=[ 
                Text("Filter Laporan Transaksi", size=14, weight=FontWeight.BOLD),
                Text("Masukkan Tanggal Awal", size=12),
                Row([ 
                        inputan_tanggal_awal, inputan_tanggal_awal_pick, 
                        IconButton(
                            icon=icons.CALENDAR_MONTH,
                            icon_color="black",
                            on_click=lambda _: inputan_tanggal_awal_pick.pick_date(),
                        ),
                    ]),
                Text("Masukkan Tanggal Akhir", size=12),
                Row([ 
                        inputan_tanggal_akhir, inputan_tanggal_akhir_pick, 
                        IconButton(
                            icon=icons.CALENDAR_MONTH,
                            icon_color="black",
                            on_click=lambda _: inputan_tanggal_akhir_pick.pick_date(),
                        ),
                    ]),
                Text("Masukkan ID Pelanggan", size=12),
                inputan_id_pelanggan,
                Text("Masukkan Metode Pembayaran", size=12),
                inputan_status_transaksi,
                Text("Masukkan Jenis Pesanan", size=12),
                inputan_jenis_transaksi,
                Row(
                    controls = [
                        ElevatedButton("Tampilkan Laporan", on_click=update_laporan_transaksi),
                    ],
                    alignment = MainAxisAlignment.CENTER,  # Menyusun tombol di tengah
                    spacing = 10,  # Menambahkan jarak antara tombol
                ),
                Row(
                    controls = [
                        ElevatedButton("Simpan Excel", icon=icons.PRINT, on_click=simpan_excel),  # Tombol untuk menyimpan ke Excel
                        ElevatedButton(
                            "Cetak PDF",
                            icon = icons.PRINT,
                            on_click = pratinjau_pdf
                        ),
                    ],
                    alignment = MainAxisAlignment.CENTER,  # Menyusun tombol di tengah
                    spacing = 10,  # Menambahkan jarak antara tombol
                ),
            ],
            width=300,
        ),
        border_radius=20,
        border=border.all(color=colors.GREY_900, width=0.5),
        padding=20,
    )

    # Menampilkan halaman laporan transaksi
    return Container(  # Wrap seluruh Column ke dalam Container
        content = Column(
            controls = [
                # Baris pertama
                Row([
                        Icon(name = icons.DATA_EXPLORATION_ROUNDED, size = 50, color = colors.BLUE_400),
                        Text("Kelola Laporan Transaksi", size = 30, weight = "bold")
                    ], alignment = MainAxisAlignment.CENTER),
                Row(
                    controls = [form_filter, form_kanan],
                    alignment = MainAxisAlignment.START,  # Menyusun item di kiri
                    vertical_alignment = CrossAxisAlignment.START,  # Menyusun item di atas secara vertikal
                ),
                snack_bar_berhasil, snack_bar_gagal
            ]
        ),
        margin = 10    # Memberikan jarak di luar container
    )
